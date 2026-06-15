# -*- coding: utf-8 -*-
"""
Barcha studentlarga ma'lum (tasodifiy) parol o'rnatadi va login+parol
ro'yxatini Excelga eksport qiladi.

- Paroli ALLAQACHON ma'lum (plain_password to'la) studentlarga TEGMAYDI
  → bir necha marta ishga tushirish xavfsiz (idempotent).
- Yangi tasodifiy parol: 8 belgi, chalkash belgilarsiz (0/O, 1/l/I yo'q).
- Natija: student_logins.xlsx (shu papkada) + Settings > Users'da ko'rinadi.

Ishga tushirish (loyiha papkasida):
    .venv\\Scripts\\python.exe export_student_logins.py
"""
import os
import secrets
import django
from datetime import datetime

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'wallstreet.settings')
django.setup()

from students.models import Student  # noqa: E402

# Chalkash belgilarsiz alfavit (telefonda terish oson)
ALPHABET = 'ABCDEFGHJKLMNPQRSTUVWXYZabcdefghijkmnpqrstuvwxyz23456789'


def gen_password(length=8):
    return ''.join(secrets.choice(ALPHABET) for _ in range(length))


def main():
    reset_count = 0
    rows = []

    students = (
        Student.objects.select_related('user')
        .order_by('first_name', 'last_name')
    )

    for s in students:
        user = s.user
        if not user:
            # Login bog'lanmagan student — Excelda belgilab qo'yamiz
            rows.append((s.full_name, s.phone, '— login yo\'q —', '—',
                         'Nofaol' if not s.is_active else 'Faol'))
            continue

        pw = (user.plain_password or '').strip()
        if not pw:
            pw = gen_password()
            user.set_password(pw)
            user.plain_password = pw
            if not user.is_active:
                user.is_active = True
            user.save(update_fields=['password', 'plain_password', 'is_active'])
            reset_count += 1

        rows.append((
            s.full_name,
            s.phone,
            user.username,
            pw,
            'Faol' if s.is_active else 'Nofaol',
        ))

    _export_excel(rows)
    print(f'YANGI_PAROL_BERILDI: {reset_count}')
    print(f'JAMI_STUDENT: {len(rows)}')


def _export_excel(rows):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        # openpyxl bo'lmasa — CSV zaxira
        import csv
        path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            'student_logins.csv')
        with open(path, 'w', newline='', encoding='utf-8-sig') as f:
            w = csv.writer(f)
            w.writerow(['#', 'F.I.Sh', 'Telefon', 'Login', 'Parol', 'Holat'])
            for i, r in enumerate(rows, 1):
                w.writerow([i, *r])
        print(f'FAYL: {path}')
        return

    wb = Workbook()
    ws = wb.active
    ws.title = 'Studentlar'

    title = f"Wall Street — Student loginlari ({datetime.now():%d.%m.%Y %H:%M})"
    ws.merge_cells('A1:F1')
    c = ws['A1']
    c.value = title
    c.font = Font(bold=True, size=13, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='0A1628')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 26

    headers = ['#', 'F.I.Sh', 'Telefon', 'Login (username)', 'Parol', 'Holat']
    ws.append(headers)
    hfill = PatternFill('solid', fgColor='1A3A5C')
    thin = Side(style='thin', color='D0D7E2')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, 7):
        cell = ws.cell(row=2, column=col)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    for i, r in enumerate(rows, 1):
        ws.append([i, *r])
        row_idx = i + 2
        for col in range(1, 7):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = border
            if col in (1, 6):
                cell.alignment = Alignment(horizontal='center')
            if col == 5:  # parol — monospace ko'rinish uchun qalin
                cell.font = Font(name='Consolas', bold=True, color='B45309')

    widths = [5, 26, 18, 22, 16, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws.freeze_panes = 'A3'

    path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        'student_logins.xlsx')
    wb.save(path)
    print(f'FAYL: {path}')


if __name__ == '__main__':
    main()
