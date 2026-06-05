import io
from datetime import datetime

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum
from django.http import HttpResponse
from django.utils import timezone

from .models import SalaryPayment, MONTH_NAMES
from .forms import SalaryPaymentForm
from instructors.models import Instructor


def _filtered_payments(request):
    """Salary list bilan bir xil filtr (year) qo'llangan queryset."""
    payments = SalaryPayment.objects.select_related('instructor').all()
    year = request.GET.get('year')
    if year:
        payments = payments.filter(year=year)
    return payments, year


def _list_context(request, extra=None):
    """Shared context for the salary list page (incl. modal form data)."""
    payments = SalaryPayment.objects.select_related('instructor').all()

    year = request.GET.get('year')
    if year:
        payments = payments.filter(year=year)

    total_paid = payments.filter(status='paid').aggregate(t=Sum('amount'))['t'] or 0
    total_pending = payments.filter(status='pending').aggregate(t=Sum('amount'))['t'] or 0

    years = (
        SalaryPayment.objects.values_list('year', flat=True)
        .distinct().order_by('-year')
    )

    ctx = {
        'payments': payments,
        'total_paid': total_paid,
        'total_pending': total_pending,
        'total_count': payments.count(),
        'years': years,
        'selected_year': int(year) if year else None,
        'instructor_count': Instructor.objects.count(),
        # modal form data
        'instructors': Instructor.objects.all(),
        'months': list(enumerate(MONTH_NAMES, start=1)),
        'statuses': SalaryPayment.STATUS_CHOICES,
        'current_year': timezone.now().year,
        'current_month': timezone.now().month,
    }
    if extra:
        ctx.update(extra)
    return ctx


@login_required
def salary_list(request):
    return render(request, 'staff_salary/salary_list.html', _list_context(request))


@login_required
def salary_create(request):
    if request.method != 'POST':
        # The form lives in a modal on the list page.
        return redirect('salary_list')

    form = SalaryPaymentForm(request.POST)
    if form.is_valid():
        form.save()
        messages.success(request, "Salary payment added successfully.")
        return redirect('salary_list')

    messages.error(request, "Please correct the errors in the salary form.")
    return render(request, 'staff_salary/salary_list.html', _list_context(request, {
        'form_errors': form.errors,
        'form_data': request.POST,
        'open_modal': 'create',
    }))


@login_required
def salary_edit(request, pk):
    payment = get_object_or_404(SalaryPayment, pk=pk)
    if request.method != 'POST':
        return redirect('salary_list')

    form = SalaryPaymentForm(request.POST, instance=payment)
    if form.is_valid():
        form.save()
        messages.success(request, "Salary payment updated.")
        return redirect('salary_list')

    messages.error(request, "Please correct the errors in the salary form.")
    return render(request, 'staff_salary/salary_list.html', _list_context(request, {
        'form_errors': form.errors,
        'form_data': request.POST,
        'open_modal': 'edit',
        'edit_pk': str(pk),
    }))


@login_required
def salary_delete(request, pk):
    payment = get_object_or_404(SalaryPayment, pk=pk)
    if request.method == 'POST':
        payment.delete()
        messages.success(request, "Salary payment deleted.")
        return redirect('salary_list')
    return render(request, 'staff_salary/salary_confirm_delete.html', {'payment': payment})


# ══════════════════════════════════════════════════════════════════════════════
#  EXPORT — Excel & PDF
# ══════════════════════════════════════════════════════════════════════════════

@login_required
def salary_export_excel(request):
    """Oylik to'lov hisobotini .xlsx ko'rinishda yuklab beradi."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    payments, year = _filtered_payments(request)
    payments = payments.order_by('-year', '-month', 'instructor__full_name')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Salaries"

    PRIMARY = "0E5C60"
    GOLD = "C9A84C"

    # ── Sarlavha ──
    title = f"Wall Street — Salary report{f' ({year})' if year else ''}"
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = title
    cell.font = Font(size=15, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=PRIMARY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:F2')
    sub = ws['A2']
    sub.value = f"Generated: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    sub.font = Font(size=9, italic=True, color="666666")
    sub.alignment = Alignment(horizontal="center")

    # ── Ustun sarlavhalari ──
    headers = ["#", "Teacher", "Amount (UZS)", "Period", "Status", "Paid date"]
    header_row = 4
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor=GOLD)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    # ── Qatorlar ──
    total_paid = 0
    total_pending = 0
    r = header_row + 1
    for i, p in enumerate(payments, start=1):
        amount = float(p.amount)
        if p.status == 'paid':
            total_paid += amount
        else:
            total_pending += amount
        values = [
            i,
            p.instructor.full_name,
            amount,
            f"{p.get_month_display()} {p.year}",
            p.get_status_display(),
            p.paid_at.strftime('%d.%m.%Y') if p.paid_at else '—',
        ]
        for col, v in enumerate(values, start=1):
            c = ws.cell(row=r, column=col, value=v)
            c.border = border
            if col == 3:
                c.number_format = '#,##0'
                c.alignment = Alignment(horizontal="right")
            elif col in (1, 4, 5, 6):
                c.alignment = Alignment(horizontal="center")
        r += 1

    # ── Jami ──
    ws.cell(row=r, column=2, value="TOTAL PAID").font = Font(bold=True)
    tp = ws.cell(row=r, column=3, value=total_paid)
    tp.font = Font(bold=True, color=PRIMARY)
    tp.number_format = '#,##0'
    tp.alignment = Alignment(horizontal="right")
    r += 1
    ws.cell(row=r, column=2, value="TOTAL PENDING").font = Font(bold=True)
    tn = ws.cell(row=r, column=3, value=total_pending)
    tn.font = Font(bold=True, color="D97706")
    tn.number_format = '#,##0'
    tn.alignment = Alignment(horizontal="right")

    # ── Ustun kengligi ──
    widths = [6, 30, 18, 20, 14, 14]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    fname = f"salary_report{f'_{year}' if year else ''}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        bio,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


@login_required
def salary_export_pdf(request):
    """Oylik to'lov hisobotini chiroyli PDF ko'rinishda yuklab beradi."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.colors import HexColor
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT

    payments, year = _filtered_payments(request)
    payments = payments.order_by('-year', '-month', 'instructor__full_name')

    PRIMARY = HexColor('#0E5C60')
    PRIMARY_DARK = HexColor('#093D40')
    GOLD = HexColor('#C9A84C')
    LIGHT = HexColor('#e6f4f5')

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=16*mm, rightMargin=16*mm,
        topMargin=16*mm, bottomMargin=16*mm,
        title="Salary report",
    )

    styles = getSampleStyleSheet()
    h_title = ParagraphStyle('wsTitle', parent=styles['Title'],
                             fontSize=20, textColor=PRIMARY_DARK, spaceAfter=2)
    h_sub = ParagraphStyle('wsSub', parent=styles['Normal'],
                           fontSize=9, textColor=colors.grey, spaceAfter=14)
    cell_style = ParagraphStyle('wsCell', parent=styles['Normal'], fontSize=9)

    elements = []
    elements.append(Paragraph("Wall Street Education Centre", h_title))
    period = f"Year {year}" if year else "All years"
    elements.append(Paragraph(
        f"Salary report &nbsp;•&nbsp; {period} &nbsp;•&nbsp; "
        f"Generated {datetime.now().strftime('%d.%m.%Y %H:%M')}", h_sub))

    # ── Jadval ma'lumotlari ──
    data = [["#", "Teacher", "Amount (UZS)", "Period", "Status", "Paid date"]]
    total_paid = 0
    total_pending = 0
    for i, p in enumerate(payments, start=1):
        amount = float(p.amount)
        if p.status == 'paid':
            total_paid += amount
        else:
            total_pending += amount
        data.append([
            str(i),
            Paragraph(p.instructor.full_name, cell_style),
            f"{amount:,.0f}",
            f"{p.get_month_display()} {p.year}",
            p.get_status_display(),
            p.paid_at.strftime('%d.%m.%Y') if p.paid_at else '—',
        ])

    if len(data) == 1:
        data.append(["—", "No salary payments", "—", "—", "—", "—"])

    table = Table(
        data,
        colWidths=[12*mm, 52*mm, 30*mm, 30*mm, 24*mm, 26*mm],
        repeatRows=1,
    )
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), PRIMARY),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9.5),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('ALIGN', (3, 0), (5, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT]),
        ('LINEBELOW', (0, 0), (-1, -1), 0.4, HexColor('#cfe3e4')),
        ('BOX', (0, 0), (-1, -1), 0.6, HexColor('#cfe3e4')),
    ])
    table.setStyle(style)
    elements.append(table)
    elements.append(Spacer(1, 12))

    # ── Jami qatorlari ──
    totals = Table([
        ["Total paid:", f"{total_paid:,.0f} UZS"],
        ["Total pending:", f"{total_pending:,.0f} UZS"],
    ], colWidths=[40*mm, 50*mm], hAlign='RIGHT')
    totals.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('TEXTCOLOR', (1, 0), (1, 0), PRIMARY_DARK),
        ('TEXTCOLOR', (1, 1), (1, 1), HexColor('#d97706')),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(totals)

    def _footer(canvas, doc_):
        canvas.saveState()
        canvas.setFillColor(PRIMARY_DARK)
        canvas.rect(0, 0, A4[0], 12*mm, fill=1, stroke=0)
        canvas.setFillColor(GOLD)
        canvas.rect(0, 12*mm, A4[0], 0.8*mm, fill=1, stroke=0)
        canvas.setFillColor(colors.white)
        canvas.setFont('Helvetica', 8)
        canvas.drawCentredString(
            A4[0] / 2, 4.5*mm,
            'Wall Street Education Centre  •  Tashkent, Uzbekistan')
        canvas.setFont('Helvetica', 7)
        canvas.drawRightString(A4[0] - 16*mm, 4.5*mm, f"Page {doc_.page}")
        canvas.restoreState()

    doc.build(elements, onFirstPage=_footer, onLaterPages=_footer)
    buffer.seek(0)

    fname = f"salary_report{f'_{year}' if year else ''}_{datetime.now().strftime('%Y%m%d')}.pdf"
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response
